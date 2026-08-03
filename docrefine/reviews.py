# SAVE AS: docrefine/reviews.py
"""
Where rebrand review sheets live.

Up to v138 the sheet was written as `_rebrand_plan.csv` *inside* the analyzed
source folder, where it was easy to lose among thousands of subfolders — and
where a "complete set" copy would sweep it into the upload tree.

From v139 every sheet goes to one fixed, obvious place:

    Documents/DocRefinePro_Data/Rebrand Reviews/

named after the source it describes, so Apply can find it again on its own.
The name carries the parent folder too, because deduplicated jobs all call their
masters folder `01_Master_Files` — the job name is what tells them apart.

Sheets written by older versions are still found (see `plan_candidates`).
"""
import csv
import re
from pathlib import Path

from .config import REVIEWS_ROOT

PLAN_STEM_SUFFIX = "_rebrand_plan"
PLAN_EXTS = (".xlsx", ".csv")          # preferred first
PLAN_SUFFIX = PLAN_STEM_SUFFIX + ".csv"   # back-compat alias
LEGACY_PLAN_NAME = "_rebrand_plan.csv"


def is_plan_file(name):
    """True for any review sheet, whatever its format or where it was written."""
    p = Path(str(name))
    n = p.name.lower()
    if n == LEGACY_PLAN_NAME:
        return True
    return p.stem.lower().endswith(PLAN_STEM_SUFFIX) and p.suffix.lower() in PLAN_EXTS


def _safe(part):
    """Make one path component safe (and short enough) for a filename."""
    cleaned = re.sub(r'[<>:"/\\|?*]+', "-", (part or "").strip())
    return cleaned.strip(". ")[:60] or "folder"


def plan_stem_for(src):
    """Stable sheet name (no extension) for a source folder: `<parent>__<folder>_rebrand_plan`."""
    try:
        p = Path(src).resolve()
    except OSError:
        p = Path(src)
    parent = p.parent.name
    stem = f"{_safe(parent)}__{_safe(p.name)}" if parent else _safe(p.name)
    return stem + PLAN_STEM_SUFFIX


def plan_name_for(src, ext=".xlsx"):
    """Stable sheet filename for a source folder."""
    return plan_stem_for(src) + ext


def plan_path_for(src, ext=".xlsx"):
    """Where to WRITE the review sheet for a source folder (creates the folder)."""
    REVIEWS_ROOT.mkdir(parents=True, exist_ok=True)
    return REVIEWS_ROOT / plan_name_for(src, ext)


def plan_candidates(src):
    """Every place a sheet for this source may live, best first."""
    p = Path(src)
    stem = plan_stem_for(src)
    out = []
    for ext in PLAN_EXTS:
        out.append(REVIEWS_ROOT / (stem + ext))          # current default
    for ext in PLAN_EXTS:
        out.append(p.parent / (stem + ext))              # hand-placed beside the source
        out.append(p.parent / f"{p.name}{PLAN_STEM_SUFFIX}{ext}")
    out.append(p / LEGACY_PLAN_NAME)                     # pre-v139: inside the source
    return out


def find_plan(src):
    """Locate an existing review sheet for a source folder, or None."""
    for c in plan_candidates(src):
        try:
            if c.is_file():
                return c
        except OSError:
            continue
    return None


# --------------------------------------------------------------------------
#  Reading and writing the sheet
# --------------------------------------------------------------------------

def read_plan(path):
    """Read a review sheet (.xlsx or .csv) into a list of dicts."""
    path = Path(path)
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            header = [str(h or "").strip() for h in next(it, ())]
            rows = []
            for raw in it:
                if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                    continue
                rows.append({h: ("" if v is None else str(v).strip())
                             for h, v in zip(header, raw) if h})
            return rows
        finally:
            wb.close()
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def needs_review(row):
    """True when a row deserves a human's eyes before Apply runs."""
    if (row.get("source") or "").strip().lower() != "llm":
        return True
    try:
        return float(row.get("confidence") or 0) < 0.9
    except (TypeError, ValueError):
        return True


def write_plan(path, rows, columns, src_root=None, asset_types=()):
    """Write the review sheet. .xlsx gets the review affordances; .csv stays plain."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() != ".xlsx":
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
            w.writeheader()
            for r in rows:
                w.writerow(r)
        return path
    _write_xlsx(path, rows, columns, src_root, asset_types)
    return path


def _write_xlsx(path, rows, columns, src_root, asset_types):
    """A sheet you can actually review: openable files, dropdowns, triage flag."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    header = ["review?"] + list(columns)
    wb = Workbook()
    ws = wb.active
    ws.title = "Rebrand plan"
    ws.append(header)

    head_fill = PatternFill("solid", fgColor="1F2A5A")
    for i, name in enumerate(header, start=1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(vertical="center")

    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    file_col = header.index("file") + 1
    for r in rows:
        ws.append(["YES" if needs_review(r) else ""] + [r.get(c, "") for c in columns])
        i = ws.max_row
        if ws.cell(row=i, column=1).value == "YES":
            ws.cell(row=i, column=1).fill = flag_fill
        # the filename opens the actual PDF — reviewing without seeing it is guesswork
        if src_root:
            target = Path(src_root) / str(r.get("file", ""))
            cell = ws.cell(row=i, column=file_col)
            try:
                cell.hyperlink = target.as_uri()
                cell.font = Font(color="0563C1", underline="single")
            except (ValueError, OSError):
                pass

    widths = {"review?": 9, "file": 52, "action": 10, "doc_type": 22, "product": 24,
              "asset_type": 22, "manufacturer": 26, "title": 30, "pages": 7,
              "confidence": 11, "source": 9, "notes": 40}
    for i, name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 16)

    ws.freeze_panes = "A2"
    last = max(ws.max_row, 2)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{last}"

    def add_list(col_name, values):
        if col_name not in header or not values:
            return
        col = get_column_letter(header.index(col_name) + 1)
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(values), allow_blank=True)
        dv.error = "Pick one of the listed values."
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{last}")

    add_list("action", ["rebrand", "leave"])
    # Excel caps an inline list at 255 characters; keep the common ones.
    picks, total = [], 0
    for a in asset_types:
        if total + len(a) + 1 > 250:
            break
        picks.append(a); total += len(a) + 1
    add_list("asset_type", picks)

    wb.save(path)
